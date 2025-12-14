from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import VehicleEntry, Cars
from .utils import print_stats_receipt, _resolve_printer_name
import sys
from django.views import View
from django.contrib.auth import login, logout, authenticate
from django.shortcuts import render, redirect
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
import json
from datetime import datetime, timedelta
import re
from django.core.files.base import ContentFile
from django.db import transaction
from django.views.decorators.http import require_POST, require_GET
from django.db.models import Sum, Q
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

# MIN_TIME_BETWEEN_ENTRIES available in settings if needed elsewhere
from django.contrib.auth.decorators import login_required
from django.contrib import messages

if sys.platform.startswith("win"):
    import win32print  # type: ignore[import]
else:
    win32print = None  # type: ignore[assignment]


class LoginView(View):
    def get(self, request):
        return render(request, "login.html")

    def post(self, request):
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("home")
        else:
            return render(
                request,
                "login.html",
                {
                    "message": "Login yoki parol xato. Iltimos tekshirib qayta urinib ko`ring!"
                },
                status=401,
            )


class LogoutView(LoginRequiredMixin, View):
    def post(self, request):
        logout(request)
        return redirect("login")

    def get(self, request):
        logout(request)
        return redirect("login")


class HomeView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, "home.html")


@csrf_exempt
@require_POST
def receive_entry(request):
    try:
        with transaction.atomic():
            content_type = request.headers.get("Content-Type", "")
            body_bytes = request.body
            body_str = body_bytes.decode("utf-8", errors="ignore")

            # 1. XML'dan davlat raqamini ajratamiz (masalan <licensePlate> tagidan)
            number_plate_match = re.search(
                r"<licensePlate>(.*?)</licensePlate>", body_str
            )
            number_plate = (
                number_plate_match.group(1)
                if number_plate_match
                else f"TEMP{timezone.now().strftime('%H%M%S')}"
            )

            # 2. Faylni multipart dan ajratish
            # Bu oddiy variant: multipart so'rovdan rasmni olish
            boundary = (
                content_type.split("boundary=")[-1]
                if "boundary=" in content_type
                else None
            )

            if boundary:
                parts = body_bytes.split(boundary.encode())
                image_data = None
                for part in parts:
                    if b"Content-Type: image/jpeg" in part:
                        image_data = part.split(b"\r\n\r\n", 1)[-1].rsplit(b"\r\n", 1)[
                            0
                        ]
                        break
                car = Cars.objects.filter(
                    number_plate=number_plate, is_blocked=True
                ).first()
                if car:
                    channel_layer = get_channel_layer()
                    async_to_sync(channel_layer.group_send)(
                        "home_updates",
                        {
                            "type": "broadcast_notification",
                            "title": "🚫 Bloklangan avtomobil",
                            "message": f"Avtomobil {number_plate} bloklangan! Chiqish taqiqlanadi.",
                            "notification_type": "error",
                            "timestamp": timezone.now().isoformat(),
                        },
                    )

                    return JsonResponse(
                        {
                            "status": "error",
                            "message": f"Bu avtomobilga taqiq qo'shilgan! {number_plate}",
                        }
                    )
                if image_data:
                    # 3. Fayl nomi: Raqam + sana (20250717_135501.jpg)
                    current_time = timezone.now()
                    timestamp = current_time.strftime("%Y%m%d_%H%M%S")
                    filename = f"{number_plate}_{timestamp}.jpg"

                    # 4. Rasmdan ImageField fayl obyektini yasaymiz
                    image_file = ContentFile(image_data, name=filename)

                    # 5. Bazaga yozamiz - entry_time auto_now_add=True bo'lgani uchun o'rnatmaymiz
                    if VehicleEntry.objects.filter(
                        number_plate=number_plate, exit_time__isnull=True
                    ).exists():
                        latest_entry = VehicleEntry.objects.filter(
                            number_plate=number_plate, exit_time__isnull=True
                        ).last()
                        latest_entry.delete()
                        VehicleEntry.objects.create(
                            number_plate=number_plate,
                            entry_image=image_file,
                            total_amount=0,
                        )

                        return JsonResponse(
                            {
                                "status": "ok",
                                "message": f"Bu avtomobilga taqiq qo'shilgan! {number_plate}",
                            }
                        )
                    entry = VehicleEntry.objects.create(
                        number_plate=number_plate,
                        entry_image=image_file,
                        total_amount=0,
                    )

                    return JsonResponse(
                        {
                            "status": "ok",
                            "message": "VehicleEntry created",
                            "number_plate": number_plate,
                            "file_saved": filename,
                            "entry_id": entry.id,
                        }
                    )
                else:
                    # Agar rasm topilmasa, xatolik xabarini qaytaramiz
                    return JsonResponse(
                        {
                            "status": "ok",
                            "message": f"Avtomobil {number_plate} uchun rasm topilmadi. Iltimos qayta urinib ko'ring.",
                        },
                        status=200,
                    )
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@require_POST
def entry_check(request):
    try:
        # Request dan nima kelishidan qat'iy nazar, oxirgi kirishni olamiz
        last_entry = VehicleEntry.objects.order_by("-entry_time").first()

        if not last_entry:
            return JsonResponse(
                {"status": "error", "message": "Hech qanday kirish topilmadi"},
                status=404,
            )

        if not last_entry.is_checked:
            # is_checked ni True qilamiz
            last_entry.is_checked = True
            last_entry.save(update_fields=["is_checked"])

            return JsonResponse(
                {
                    "status": "ok",
                    "id": last_entry.id,
                    "uuid": str(last_entry.uuid),
                    "license_plate": last_entry.number_plate,
                    "entry_time": last_entry.entry_time.isoformat(),
                }
            )
        else:
            # Yangi kirish yaratish
            new_entry = default_create_entry(request)

            if new_entry:
                return JsonResponse(
                    {
                        "status": "ok",
                        "id": new_entry.id,
                        "uuid": str(new_entry.uuid),
                        "license_plate": new_entry.number_plate,
                        "entry_time": new_entry.entry_time.isoformat(),
                    }
                )
            else:
                return JsonResponse(
                    {"status": "error", "message": "Yangi kirish yaratishda xatolik"},
                    status=500,
                )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def default_create_entry(request):
    """
    Default kirish yozuvi yaratish (rasm bilan)
    """
    try:
        with transaction.atomic():
            from django.core.files.base import ContentFile
            from PIL import Image
            import io

            img = Image.new("RGB", (1, 1), color="white")
            img_io = io.BytesIO()
            img.save(img_io, format="JPEG")
            img_file = ContentFile(img_io.getvalue(), name="default_entry.jpg")

            new_entry = VehicleEntry.objects.create(
                number_plate="NOMA'LUM",
                entry_image=img_file,
                total_amount=0,
            )
            return new_entry
    except Exception as e:
        print(f"default_create_entry xatolik: {str(e)}")
        return None


@csrf_exempt
@require_POST
def receive_exit(request):
    try:
        with transaction.atomic():
            content_type = request.headers.get("Content-Type", "")
            body_bytes = request.body
            body_str = body_bytes.decode("utf-8", errors="ignore")
            current_time = timezone.now()

            # Create timezone-aware datetime range for today
            today = current_time.date()
            start_datetime = timezone.make_aware(
                datetime.combine(today, datetime.min.time())
            )
            end_datetime = timezone.make_aware(
                datetime.combine(today, datetime.max.time())
            )

            # 1. XML'dan davlat raqamini ajratamiz (masalan <licensePlate> tagidan)
            number_plate_match = re.search(
                r"<licensePlate>(.*?)</licensePlate>", body_str
            )
            number_plate = (
                number_plate_match.group(1)
                if number_plate_match
                else f"TEMP{current_time.strftime('%H%M%S')}"
            )

            car = Cars.objects.filter(number_plate=number_plate).first()

            # 2. Faylni multipart dan ajratish
            # Bu oddiy variant: multipart so'rovdan rasmni olish
            boundary = (
                content_type.split("boundary=")[-1]
                if "boundary=" in content_type
                else None
            )

            if not boundary:
                # Try to find image data without boundary
                if b"Content-Type: image/jpeg" in body_bytes:
                    # Simple approach: look for JPEG data
                    jpeg_start = body_bytes.find(b"\xff\xd8\xff")  # JPEG start marker
                    if jpeg_start != -1:
                        image_data = body_bytes[jpeg_start:]
                    else:
                        return JsonResponse(
                            {
                                "status": "error",
                                "message": "Rasm topilmadi",
                            },
                            status=400,
                        )
                else:
                    return JsonResponse(
                        {
                            "status": "error",
                            "message": "Multipart boundary topilmadi",
                        },
                        status=400,
                    )
            else:
                parts = body_bytes.split(boundary.encode())
                image_data = None
                for i, part in enumerate(parts):
                    if b"Content-Type: image/jpeg" in part:
                        image_data = part.split(b"\r\n\r\n", 1)[-1].rsplit(b"\r\n", 1)[
                            0
                        ]
                        break

            # If no image data found, we'll still process the exit
            # This handles cases where only XML data is sent
            if not image_data:
                # Create a placeholder image file or skip image processing
                image_file = None
            else:
                # 3. Fayl nomi: Raqam + sana (20250717_135501.jpg)
                timestamp = current_time.strftime("%Y%m%d_%H%M%S")
                filename = f"{number_plate}_{timestamp}.jpg"
                # 4. Rasmdan ImageField fayl obyektini yasaymiz
                image_file = ContentFile(image_data, name=filename)

            # Use timezone-aware datetime range instead of naive date

            # First, prefer the OLDEST open entry (edited/backdated entries should be used)
            latest_entry = (
                VehicleEntry.objects.filter(
                    number_plate=number_plate,
                    exit_time__isnull=True,
                )
                .order_by("entry_time")
                .first()
            )

            if not latest_entry:
                # Fall back to the most recent entry if no open entry exists
                latest_entry = (
                    VehicleEntry.objects.filter(number_plate=number_plate)
                    .order_by("-entry_time")
                    .first()
                )
                if not latest_entry:
                    # Instead of returning 404, we'll create a temporary entry and continue
                    # This allows the exit to proceed even if no entry was found
                    latest_entry = None

            # If no entry was found, we'll just process the exit without updating a database entry
            if latest_entry is None:
                # Send notification that vehicle exited but no entry was found
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    "home_updates",
                    {
                        "type": "broadcast_notification",
                        "title": "⚠️ Avtomobil chiqib ketdi",
                        "message": f"Avtomobil {number_plate} chiqib ketdi (kirish yozuvi topilmadi)",
                        "notification_type": "warning",
                        "timestamp": timezone.now().isoformat(),
                    },
                )

                # Return success response even without database entry
                return JsonResponse(
                    {
                        "status": "ok",
                        "number_plate": number_plate,
                        "amount": 0,
                        "message": "Exit processed (no entry found)",
                    }
                )

            # Check if vehicle has already exited
            if latest_entry.exit_time:
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    "home_updates",
                    {
                        "type": "broadcast_notification",
                        "title": "🚫 Avtomobil allaqachon chiqib ketgan",
                        "message": f"Avtomobil {number_plate} allaqachon chiqib ketgan!",
                        "notification_type": "warning",
                        "timestamp": timezone.now().isoformat(),
                    },
                )
                return JsonResponse(
                    {
                        "status": "error",
                        "message": f"Avtomobil {number_plate} allaqachon chiqib ketgan!",
                    }
                )

            # Check if car is blocked
            if car and car.is_blocked:
                # Send real-time notification about blocked car
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    "home_updates",
                    {
                        "type": "broadcast_notification",
                        "title": "🚫 Bloklangan avtomobil",
                        "message": f"Avtomobil {number_plate} bloklangan! Chiqish taqiqlanadi.",
                        "notification_type": "error",
                        "timestamp": timezone.now().isoformat(),
                    },
                )

                return JsonResponse(
                    {
                        "status": "error",
                        "message": f"Bu avtomobilga taqiq qo'shilgan! {number_plate}",
                    }
                )

            # Process the exit based on car type
            if car and car.is_free and not car.is_blocked:
                if image_file:
                    latest_entry.exit_image = image_file
                latest_entry.total_amount = 0
                latest_entry.exit_time = current_time
                latest_entry.save()
            elif car and car.is_special_taxi and not car.is_blocked:
                if image_file:
                    latest_entry.exit_image = image_file
                latest_entry.total_amount = (
                    latest_entry.calculate_amount()
                    if VehicleEntry.objects.filter(
                        entry_time__gte=start_datetime,
                        entry_time__lte=end_datetime,
                        number_plate=number_plate,
                    ).count()
                    < 2
                    else 0
                )
                latest_entry.exit_time = current_time
                latest_entry.save()
            else:
                if (car and not car.is_blocked) or not car:
                    if image_file:
                        latest_entry.exit_image = image_file
                    latest_entry.exit_time = current_time
                    latest_entry.total_amount = latest_entry.calculate_amount()
                    latest_entry.save()

            return JsonResponse(
                {
                    "status": "ok",
                    "number_plate": number_plate,
                    "amount": latest_entry.total_amount,
                }
            )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# New API endpoints for WebSocket functionality


@csrf_exempt
@login_required
@require_GET
def get_statistics(request):
    """Get statistics for a specific date"""
    date_str = request.GET.get("date", timezone.now().date().isoformat())
    try:
        # Convert date string to timezone-aware datetime for proper filtering
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        # Create timezone-aware datetime range for the entire day
        start_datetime = timezone.make_aware(
            datetime.combine(date_obj, datetime.min.time())
        )
        end_datetime = timezone.make_aware(
            datetime.combine(date_obj, datetime.max.time())
        )
    except ValueError:
        # Fallback to today if date parsing fails
        today = timezone.now().date()
        start_datetime = timezone.make_aware(
            datetime.combine(today, datetime.min.time())
        )
        end_datetime = timezone.make_aware(datetime.combine(today, datetime.max.time()))

    # Get entries that either:
    # 1. Entered today (normal case)
    # 2. Entered yesterday but exited today (overnight parking)
    today_entries = VehicleEntry.objects.filter(
        entry_time__gte=start_datetime, entry_time__lte=end_datetime
    )

    # Also include vehicles that entered yesterday but exited today
    yesterday_start = timezone.make_aware(
        datetime.combine(date_obj - timezone.timedelta(days=1), datetime.min.time())
    )
    yesterday_end = timezone.make_aware(
        datetime.combine(date_obj - timezone.timedelta(days=1), datetime.max.time())
    )

    overnight_entries = VehicleEntry.objects.filter(
        entry_time__gte=yesterday_start,
        entry_time__lte=yesterday_end,
        exit_time__gte=start_datetime,
        exit_time__lte=end_datetime,
    )

    # Calculate statistics separately for each queryset
    today_total = today_entries.count()
    today_exits = today_entries.filter(exit_time__isnull=False).count()
    today_inside = today_entries.filter(exit_time__isnull=True).count()
    today_unpaid = today_entries.filter(is_paid=False, exit_time__isnull=False).count()

    overnight_total = overnight_entries.count()
    overnight_exits = overnight_entries.filter(exit_time__isnull=False).count()
    overnight_inside = overnight_entries.filter(exit_time__isnull=True).count()
    overnight_unpaid = overnight_entries.filter(
        is_paid=False, exit_time__isnull=False
    ).count()

    # Combine the results
    total_entries = today_total + overnight_total
    total_exits = today_exits + overnight_exits
    total_inside = today_inside + overnight_inside
    unpaid_entries = today_unpaid + overnight_unpaid

    return JsonResponse(
        {
            "total_entries": total_entries,
            "total_exits": total_exits,
            "total_inside": total_inside,
            "unpaid_entries": unpaid_entries,
        }
    )


@csrf_exempt
@require_GET
@login_required
def get_vehicle_entries(request):
    """Get vehicle entries for a specific date with filters"""
    try:
        date_str = request.GET.get("date", timezone.now().date().isoformat())
        number_plate_filter = request.GET.get("number_plate", "")
        status_filter = request.GET.get(
            "status", "all"
        )  # all, paid, unpaid, inside, exited

        try:
            # Convert date string to timezone-aware datetime for proper filtering
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            # Create timezone-aware datetime range for the entire day
            start_datetime = timezone.make_aware(
                datetime.combine(date_obj, datetime.min.time())
            )
            end_datetime = timezone.make_aware(
                datetime.combine(date_obj, datetime.max.time())
            )
        except ValueError:
            # Fallback to today if date parsing fails
            today = timezone.now().date()
            start_datetime = timezone.make_aware(
                datetime.combine(today, datetime.min.time())
            )
            end_datetime = timezone.make_aware(
                datetime.combine(today, datetime.max.time())
            )

        # Get entries that either:
        # 1. Entered today (normal case)
        # 2. Entered yesterday but exited today (overnight parking)
        today_entries = VehicleEntry.objects.filter(
            entry_time__gte=start_datetime, entry_time__lte=end_datetime
        )

        # Also include vehicles that entered yesterday but exited today
        yesterday_start = timezone.make_aware(
            datetime.combine(date_obj - timezone.timedelta(days=1), datetime.min.time())
        )
        yesterday_end = timezone.make_aware(
            datetime.combine(date_obj - timezone.timedelta(days=1), datetime.max.time())
        )

        overnight_entries = VehicleEntry.objects.filter(
            entry_time__gte=yesterday_start,
            entry_time__lte=yesterday_end,
            exit_time__gte=start_datetime,
            exit_time__lte=end_datetime,
        )

        # Get all entry IDs from both querysets
        today_ids = list(today_entries.values_list("id", flat=True))
        overnight_ids = list(overnight_entries.values_list("id", flat=True))
        all_ids = today_ids + overnight_ids

        # Get all entries by IDs and order by entry_time
        entries = VehicleEntry.objects.filter(id__in=all_ids).order_by("-entry_time")

        if number_plate_filter:
            entries = entries.filter(number_plate__icontains=number_plate_filter)

        # Status filter
        if status_filter == "paid":
            entries = entries.filter(is_paid=True)
        elif status_filter == "unpaid":
            entries = entries.filter(is_paid=False, exit_time__isnull=False)
        elif status_filter == "inside":
            entries = entries.filter(exit_time__isnull=True)
        elif status_filter == "exited":
            entries = entries.filter(exit_time__isnull=False)

        entries_data = []
        for entry in entries:
            # Ensure timezone-aware formatting
            entry_time = (
                timezone.localtime(entry.entry_time)
                if timezone.is_naive(entry.entry_time)
                else entry.entry_time
            )
            exit_time = (
                timezone.localtime(entry.exit_time)
                if entry.exit_time and timezone.is_naive(entry.exit_time)
                else entry.exit_time
            )

            entries_data.append(
                {
                    "id": entry.id,
                    "number_plate": entry.number_plate,
                    "entry_time": entry_time.strftime("%H:%M"),
                    "exit_time": exit_time.strftime("%H:%M") if exit_time else None,
                    "total_amount": entry.total_amount or 0,
                    "is_paid": entry.is_paid,
                    "entry_image": entry.entry_image.url if entry.entry_image else None,
                    "exit_image": entry.exit_image.url if entry.exit_image else None,
                    "status": "inside"
                    if not entry.exit_time
                    else ("paid" if entry.is_paid else "unpaid"),
                }
            )

        return JsonResponse({"status": "ok", "entries": entries_data})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@require_GET
@login_required
def export_entries_xls(request):
    """Avtomobil yozuvlarini tanlangan sana uchun .xls formatida eksport qiladi"""
    try:
        date_str = request.GET.get("date", timezone.now().date().isoformat())
        number_plate_filter = request.GET.get("number_plate", "")
        status_filter = request.GET.get("status", "all")

        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            start_datetime = timezone.make_aware(
                datetime.combine(date_obj, datetime.min.time())
            )
            end_datetime = timezone.make_aware(
                datetime.combine(date_obj, datetime.max.time())
            )
        except ValueError:
            today = timezone.now().date()
            start_datetime = timezone.make_aware(
                datetime.combine(today, datetime.min.time())
            )
            end_datetime = timezone.make_aware(
                datetime.combine(today, datetime.max.time())
            )

        entries = VehicleEntry.objects.filter(
            entry_time__gte=start_datetime,
            entry_time__lte=end_datetime,
        ).order_by("-entry_time")

        if number_plate_filter:
            entries = entries.filter(number_plate__icontains=number_plate_filter)

        if status_filter == "paid":
            entries = entries.filter(is_paid=True)
        elif status_filter == "unpaid":
            entries = entries.filter(is_paid=False, exit_time__isnull=False)
        elif status_filter == "inside":
            entries = entries.filter(exit_time__isnull=True)
        elif status_filter == "exited":
            entries = entries.filter(exit_time__isnull=False)

        # Excel ochishi uchun HTML jadval yaratamiz (to'liq o'zbekcha)
        rows = [
            "<tr><th>Avtomobil raqami</th><th>Kirish vaqti</th><th>Chiqish vaqti</th><th>To'lov miqdori (so'm)</th><th>Holati</th></tr>"
        ]
        for entry in entries:
            entry_time = entry.entry_time.strftime("%H:%M") if entry.entry_time else ""
            exit_time = entry.exit_time.strftime("%H:%M") if entry.exit_time else ""
            # Holatni o'zbekchaga o'tkazamiz
            if not entry.exit_time:
                status = "Ichkarida"
            else:
                status = "To'langan" if entry.is_paid else "To'lanmagan"
            rows.append(
                f"<tr>"
                f"<td>{entry.number_plate}</td>"
                f"<td>{entry_time}</td>"
                f"<td>{exit_time}</td>"
                f"<td>{entry.total_amount or 0}</td>"
                f"<td>{status}</td>"
                f"</tr>"
            )

        # Jadval uchun style qo'shamiz: barcha kataklarga chegaralar, collapse yoqilgan
        table_html = (
            "<html>"
            "<head>"
            "<meta http-equiv='Content-Type' content='application/vnd.ms-excel; charset=UTF-8'/>"
            "<meta charset='UTF-8'/>"
            "<style>"
            "table { border-collapse: collapse; }"
            "th, td { border: 1px solid #000000; padding: 4px; mso-number-format:'\@'; }"
            "th { background: #f1f5f9; }"
            "</style>"
            "</head>"
            "<body>"
            "<table>" + "".join(rows) + "</table>"
            "</body>"
            "</html>"
        )

        # UTF-8 BOM qo'shamiz (Excelda o'zbekcha harflar to'g'ri chiqishi uchun)
        content_with_bom = "\ufeff" + table_html

        response = HttpResponse(
            content_with_bom, content_type="application/vnd.ms-excel; charset=UTF-8"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=avtomobillar_{date_str}.xls"
        )
        return response
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@require_POST
@login_required
def mark_as_paid(request):
    """Mark a vehicle entry as paid"""
    try:
        data = json.loads(request.body)
        entry_id = data.get("entry_id")

        entry = VehicleEntry.objects.get(id=entry_id)
        entry.is_paid = True
        entry.save()

        return JsonResponse({"success": True, "entry_id": entry_id})
    except VehicleEntry.DoesNotExist:
        return JsonResponse({"success": False, "error": "Entry not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@csrf_exempt
@require_POST
@login_required
def add_car(request):
    """Add or update a car with boolean flags"""
    try:
        data = json.loads(request.body)
        number_plate = data.get("number_plate")
        car_type = data.get(
            "car_type", ""
        )  # "free", "special_taxi", "blocked", "normal"
        position = data.get("position", "")
        is_blocked = data.get("is_blocked", False)

        if not number_plate:
            return JsonResponse(
                {"success": False, "error": "Number plate is required"}, status=400
            )

        # Set flags based on car type
        is_free = car_type == "free"
        is_special_taxi = car_type == "special_taxi"
        is_blocked = car_type == "blocked" or is_blocked

        # Validate position for free cars
        if is_free and not position:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Bepul avtomobillar uchun lavozim kiritish majburiy",
                },
                status=400,
            )

        car, created = Cars.objects.get_or_create(
            number_plate=number_plate,
            defaults={
                "is_free": is_free,
                "is_special_taxi": is_special_taxi,
                "is_blocked": is_blocked,
                "position": position if is_free else None,
            },
        )

        if not created:
            car.is_free = is_free
            car.is_special_taxi = is_special_taxi
            car.is_blocked = is_blocked
            car.position = position if is_free else None
            car.save()

        return JsonResponse(
            {
                "success": True,
                "car": {
                    "number_plate": car.number_plate,
                    "is_free": car.is_free,
                    "is_special_taxi": car.is_special_taxi,
                    "is_blocked": car.is_blocked,
                    "position": car.position,
                },
            }
        )
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@csrf_exempt
@require_POST
@login_required
def block_car(request):
    """Block a car by number plate"""
    try:
        data = json.loads(request.body)
        number_plate = data.get("number_plate")

        if not number_plate:
            return JsonResponse(
                {"success": False, "error": "Number plate is required"}, status=400
            )

        car = Cars.objects.get(number_plate=number_plate)
        car.is_blocked = True
        car.save()

        return JsonResponse({"success": True, "number_plate": number_plate})
    except Cars.DoesNotExist:
        return JsonResponse({"success": False, "error": "Car not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@csrf_exempt
@require_GET
@login_required
def get_unpaid_entries(request):
    """Get unpaid entries for receipt printing"""
    try:
        date_str = request.GET.get("date", timezone.now().date().isoformat())

        try:
            # Convert date string to timezone-aware datetime for proper filtering
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            # Create timezone-aware datetime range for the entire day
            start_datetime = timezone.make_aware(
                datetime.combine(date_obj, datetime.min.time())
            )
            end_datetime = timezone.make_aware(
                datetime.combine(date_obj, datetime.max.time())
            )
        except ValueError:
            # Fallback to today if date parsing fails
            today = timezone.now().date()
            start_datetime = timezone.make_aware(
                datetime.combine(today, datetime.min.time())
            )
            end_datetime = timezone.make_aware(
                datetime.combine(today, datetime.max.time())
            )

        # Get unpaid entries that have exited today
        # This includes vehicles that entered yesterday but exited today
        unpaid_entries = VehicleEntry.objects.filter(
            exit_time__gte=start_datetime,
            exit_time__lte=end_datetime,
            is_paid=False,
            exit_time__isnull=False,
        ).order_by("-exit_time")

        entries_data = []
        for entry in unpaid_entries:
            # Ensure timezone-aware formatting
            entry_time = entry.entry_time
            exit_time = entry.exit_time

            entries_data.append(
                {
                    "id": entry.id,
                    "number_plate": entry.number_plate,
                    "entry_time": entry_time.strftime("%H:%M"),
                    "exit_time": exit_time.strftime("%H:%M"),
                    "total_amount": entry.total_amount or 0,
                    "duration_hours": (exit_time - entry_time).total_seconds() / 3600,
                }
            )

        return JsonResponse({"status": "ok", "entries": entries_data})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


class FreePlateNumberView(LoginRequiredMixin, View):
    def get(self, request):
        free_plates = Cars.objects.filter(is_free=True)
        return render(request, "freeplatenumber.html", {"free_plates": free_plates})

    def post(self, request):
        number_plate = request.POST.get("number_plate")

        if Cars.objects.filter(number_plate=number_plate).exists():
            return JsonResponse(
                {"success": False, "message": "Bu raqamli avtomobil allaqachon mavjud"},
                status=400,
            )
        else:
            Cars.objects.create(number_plate=number_plate, is_free=True)
            return JsonResponse(
                {"success": True, "message": "Avtomobil muvaffaqiyatli qo`shildi"},
                status=200,
            )


class DeleteFreePlateView(LoginRequiredMixin, View):
    def get(self, request, pk):
        Cars.objects.filter(pk=pk).delete()
        return JsonResponse(
            {"success": True, "message": "Avtomobil muvaffaqiyatli o`chirildi"},
            status=200,
        )


@csrf_exempt
@require_GET
@login_required
def get_receipt(request):
    """Get receipt data for a specific entry"""
    try:
        entry_id = request.GET.get("entry_id")
        if not entry_id:
            return JsonResponse({"error": "Entry ID is required"}, status=400)

        entry = VehicleEntry.objects.get(id=entry_id)

        if not entry.is_paid:
            return JsonResponse({"error": "Entry is not paid"}, status=400)

        # Ensure timezone-aware formatting
        entry_time = entry.entry_time
        exit_time = entry.exit_time

        receipt_data = {
            "id": entry.id,
            "number_plate": entry.number_plate,
            "entry_time": entry_time.strftime("%H:%M"),
            "exit_time": exit_time.strftime("%H:%M") if exit_time else None,
            "total_amount": entry.total_amount or 0,
            "is_paid": entry.is_paid,
            "duration_hours": (exit_time - entry_time).total_seconds() / 3600
            if exit_time
            else 0,
        }

        return JsonResponse({"status": "ok", "receipt": receipt_data})

    except VehicleEntry.DoesNotExist:
        return JsonResponse({"error": "Entry not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


class CarsManagementView(LoginRequiredMixin, View):
    def get(self, request):
        cars = Cars.objects.all().order_by("-id")

        # Calculate stats
        total_cars = cars.count()
        free_cars = cars.filter(is_free=True).count()
        special_taxi = cars.filter(is_special_taxi=True).count()
        blocked_cars = cars.filter(is_blocked=True).count()

        context = {
            "cars": cars,
            "total_cars": total_cars,
            "free_cars": free_cars,
            "special_taxi": special_taxi,
            "blocked_cars": blocked_cars,
        }
        return render(request, "cars_management.html", context)


@csrf_exempt
@require_POST
@login_required
def create_car(request):
    """Create a new car"""
    try:
        data = json.loads(request.body)
        number_plate = data.get("number_plate")
        car_type = data.get(
            "car_type", ""
        )  # "free", "special_taxi", "blocked", "normal"
        position = data.get("position", "")
        is_blocked = data.get("is_blocked", False)

        if not number_plate:
            return JsonResponse(
                {"success": False, "error": "Number plate is required"}, status=400
            )

        # Check if car already exists
        if Cars.objects.filter(number_plate=number_plate).exists():
            return JsonResponse(
                {"success": False, "error": "Bu raqamli avtomobil allaqachon mavjud"},
                status=400,
            )

        # Set flags based on car type
        is_free = car_type == "free"
        is_special_taxi = car_type == "special_taxi"
        is_blocked = car_type == "blocked" or is_blocked

        # Validate position for free cars
        if is_free and not position:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Bepul avtomobillar uchun lavozim kiritish majburiy",
                },
                status=400,
            )

        car = Cars.objects.create(
            number_plate=number_plate,
            is_free=is_free,
            is_special_taxi=is_special_taxi,
            is_blocked=is_blocked,
            position=position if is_free else None,
        )

        return JsonResponse(
            {
                "success": True,
                "car": {
                    "id": car.id,
                    "number_plate": car.number_plate,
                    "is_free": car.is_free,
                    "is_special_taxi": car.is_special_taxi,
                    "is_blocked": car.is_blocked,
                    "position": car.position,
                },
            }
        )
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@csrf_exempt
@require_POST
@login_required
def update_car(request, car_id):
    """Update an existing car"""
    try:
        data = json.loads(request.body)
        car_type = data.get(
            "car_type", ""
        )  # "free", "special_taxi", "blocked", "normal"
        position = data.get("position", "")
        is_blocked = data.get("is_blocked", False)

        car = Cars.objects.get(id=car_id)

        # Set flags based on car type
        car.is_free = car_type == "free"
        car.is_special_taxi = car_type == "special_taxi"
        car.is_blocked = car_type == "blocked" or is_blocked

        # Validate position for free cars
        if car_type == "free" and not position:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Bepul avtomobillar uchun lavozim kiritish majburiy",
                },
                status=400,
            )

        car.position = position if car_type == "free" else None
        car.save()

        return JsonResponse(
            {
                "success": True,
                "car": {
                    "id": car.id,
                    "number_plate": car.number_plate,
                    "is_free": car.is_free,
                    "is_special_taxi": car.is_special_taxi,
                    "is_blocked": car.is_blocked,
                    "position": car.position,
                },
            }
        )
    except Cars.DoesNotExist:
        return JsonResponse({"success": False, "error": "Car not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@csrf_exempt
@require_POST
@login_required
def delete_car(request, car_id):
    """Delete a car"""
    try:
        car = Cars.objects.get(id=car_id)
        car.delete()
        return JsonResponse({"success": True, "car_id": car_id})
    except Cars.DoesNotExist:
        return JsonResponse({"success": False, "error": "Car not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@csrf_exempt
@require_POST
@login_required
def upload_license(request):
    """Upload license file for special taxi"""
    try:
        if "license_file" not in request.FILES:
            return JsonResponse(
                {"success": False, "error": "Fayl yuklanmadi"}, status=400
            )

        license_file = request.FILES["license_file"]
        car_id = request.POST.get("car_id")

        if not car_id:
            return JsonResponse(
                {"success": False, "error": "Avtomobil ID kerak"}, status=400
            )

        car = Cars.objects.get(id=car_id)
        car.license_file = license_file
        car.save()

        return JsonResponse(
            {
                "success": True,
                "license_url": car.license_file.url if car.license_file else None,
            }
        )
    except Cars.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "Avtomobil topilmadi"}, status=404
        )
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


class UnpaidEntriesView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, "unpaid_entries.html")


class DetailedEntriesView(LoginRequiredMixin, View):
    """Detailed view for vehicle entries with advanced filtering and search"""

    def get(self, request):
        return render(request, "detailed_entries.html")


@csrf_exempt
@require_GET
@login_required
def get_detailed_entries(request):
    """Get detailed vehicle entries with advanced filtering and search"""
    try:
        # Get filter parameters
        date_from = request.GET.get("date_from", "")
        date_to = request.GET.get("date_to", "")
        number_plate_filter = request.GET.get("number_plate", "")
        payment_status = request.GET.get("payment_status", "all")  # all, paid, unpaid
        entry_status = request.GET.get("entry_status", "all")  # all, inside, exited
        search_query = request.GET.get("search", "")
        export_format = request.GET.get("export", "")

        # If no dates provided, default to today's date range
        if not date_from and not date_to:
            today_str = timezone.now().date().isoformat()
            date_from = today_str
            date_to = today_str

        # Build base queryset
        entries = VehicleEntry.objects.all()

        # Compute date window once
        start_datetime = None
        end_datetime = None
        try:
            if date_from:
                start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                start_datetime = timezone.make_aware(
                    datetime.combine(start_date, datetime.min.time())
                )
            if date_to:
                end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                end_datetime = timezone.make_aware(
                    datetime.combine(end_date, datetime.max.time())
                )
        except ValueError:
            start_datetime = None
            end_datetime = None

        # Listing logic
        if entry_status == "inside":
            # Only cars without exit, entered within window
            entries = entries.filter(exit_time__isnull=True)
            if start_datetime:
                entries = entries.filter(entry_time__gte=start_datetime)
            if end_datetime:
                entries = entries.filter(entry_time__lte=end_datetime)
        elif entry_status == "exited":
            # Only cars that exited within window
            if start_datetime:
                entries = entries.filter(exit_time__gte=start_datetime)
            if end_datetime:
                entries = entries.filter(exit_time__lte=end_datetime)
            entries = entries.filter(exit_time__isnull=False)
        else:
            # All: show cars that either entered or exited within window
            time_filter = None
            if start_datetime or end_datetime:
                time_filter = Q()
                if start_datetime:
                    time_filter &= Q(entry_time__gte=start_datetime) | Q(
                        exit_time__gte=start_datetime
                    )
                if end_datetime:
                    time_filter &= Q(entry_time__lte=end_datetime) | Q(
                        exit_time__lte=end_datetime
                    )
            if time_filter is not None:
                entries = entries.filter(time_filter)

        # Number plate filter
        if number_plate_filter:
            entries = entries.filter(number_plate__icontains=number_plate_filter)

        # Payment status filter
        if payment_status == "paid":
            entries = entries.filter(is_paid=True)
        elif payment_status == "unpaid":
            entries = entries.filter(is_paid=False)

        # Re-affirm entry status filters applied above (no-op here)

        # Search query (searches in number plate)
        if search_query:
            entries = entries.filter(number_plate__icontains=search_query)

        # Handle export
        if export_format == "xls":
            # Use date_from if available, otherwise use current date
            export_date = date_from if date_from else timezone.now().date().isoformat()
            return export_detailed_entries_xls(entries, export_date)

        # Compute summary (exit-based for revenue; exclude zero-amount from paid)
        summary_qs = VehicleEntry.objects.all()
        if start_datetime:
            summary_qs = summary_qs.filter(exit_time__gte=start_datetime)
        if end_datetime:
            summary_qs = summary_qs.filter(exit_time__lte=end_datetime)
        # Apply plate/search filters to summary as well
        if number_plate_filter:
            summary_qs = summary_qs.filter(number_plate__icontains=number_plate_filter)
        if search_query:
            summary_qs = summary_qs.filter(number_plate__icontains=search_query)
        # Compute metrics
        paid_qs = summary_qs.filter(
            is_paid=True, total_amount__gt=0, exit_time__isnull=False
        )
        paid_count = paid_qs.count()
        paid_sum = paid_qs.aggregate(total=Sum("total_amount"))["total"] or 0
        unpaid_zero_count = summary_qs.filter(is_paid=True, total_amount=0).count()
        # Inside count for today's (or selected) date only
        inside_qs = VehicleEntry.objects.filter(exit_time__isnull=True)
        try:
            if date_from:
                start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                start_datetime = timezone.make_aware(
                    datetime.combine(start_date, datetime.min.time())
                )
                inside_qs = inside_qs.filter(entry_time__gte=start_datetime)
            if date_to:
                end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                end_datetime = timezone.make_aware(
                    datetime.combine(end_date, datetime.max.time())
                )
                inside_qs = inside_qs.filter(entry_time__lte=end_datetime)
        except ValueError:
            pass
        inside_count = inside_qs.count()
        exited_count = summary_qs.filter(exit_time__isnull=False).count()

        # Pagination
        page = int(request.GET.get("page", 1))
        per_page = int(request.GET.get("per_page", 50))
        start = (page - 1) * per_page
        end = start + per_page

        # Order by most recent activity
        ordered = entries.order_by("-exit_time", "-entry_time")
        total_count = ordered.count()
        entries_page = ordered[start:end]

        # Prepare data for response
        entries_data = []
        for entry in entries_page:
            # Calculate duration if exit time exists
            duration = None
            if entry.exit_time:
                duration = entry.exit_time - entry.entry_time
                duration_hours = duration.total_seconds() / 3600
                duration_str = f"{duration_hours:.1f} soat"
            else:
                duration_str = "Ichkarida"

            # Format times
            entry_time = entry.entry_time.strftime("%Y-%m-%d %H:%M")
            exit_time = (
                entry.exit_time.strftime("%Y-%m-%d %H:%M") if entry.exit_time else None
            )

            entries_data.append(
                {
                    "id": entry.id,
                    "number_plate": entry.number_plate,
                    "entry_time": entry_time,
                    "exit_time": exit_time,
                    "duration": duration_str,
                    "total_amount": entry.total_amount or 0,
                    "is_paid": entry.is_paid,
                    "payment_status": "To'langan" if entry.is_paid else "To'lanmagan",
                    "entry_status": "Ichkarida"
                    if not entry.exit_time
                    else "Chiqib ketgan",
                    "entry_image": entry.entry_image.url if entry.entry_image else None,
                    "exit_image": entry.exit_image.url if entry.exit_image else None,
                }
            )

        return JsonResponse(
            {
                "status": "ok",
                "entries": entries_data,
                "pagination": {
                    "current_page": page,
                    "per_page": per_page,
                    "total_count": total_count,
                    "total_pages": (total_count + per_page - 1) // per_page,
                },
                "summary": {
                    "paid_count": paid_count,
                    "paid_sum": paid_sum,
                    "unpaid_zero_count": unpaid_zero_count,
                    "inside_count": inside_count,
                    "exited_count": exited_count,
                },
            }
        )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@require_GET
@login_required
def print_detailed_stats_receipt(request):
    """Print only summary statistics for detailed entries filters to local printer."""
    try:
        # Get filter parameters
        date_from = request.GET.get("date_from", "")
        date_to = request.GET.get("date_to", "")
        number_plate_filter = request.GET.get("number_plate", "")
        payment_status = request.GET.get("payment_status", "all")  # all, paid, unpaid
        entry_status = request.GET.get("entry_status", "all")  # all, inside, exited
        search_query = request.GET.get("search", "")
        # Allow overriding via query param; default to your installed printer name
        printer_name = request.GET.get("printer_name") or "XP-80 (copy 1)"

        # Default to today's date range if none provided
        if not date_from and not date_to:
            today_str = timezone.now().date().isoformat()
            date_from = today_str
            date_to = today_str

        # Base queryset (exit-based reporting)
        entries = VehicleEntry.objects.all()

        # Date range by exit_time
        if date_from:
            try:
                start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                start_datetime = timezone.make_aware(
                    datetime.combine(start_date, datetime.min.time())
                )
                entries = entries.filter(exit_time__gte=start_datetime)
            except ValueError:
                pass

        if date_to:
            try:
                end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                end_datetime = timezone.make_aware(
                    datetime.combine(end_date, datetime.max.time())
                )
                entries = entries.filter(exit_time__lte=end_datetime)
            except ValueError:
                pass

        # Number plate filter
        if number_plate_filter:
            entries = entries.filter(number_plate__icontains=number_plate_filter)

        # Payment status filter
        if payment_status == "paid":
            entries = entries.filter(is_paid=True)
        elif payment_status == "unpaid":
            entries = entries.filter(is_paid=False)

        # Entry status filter
        if entry_status == "inside":
            entries = entries.filter(exit_time__isnull=True)
        elif entry_status == "exited":
            entries = entries.filter(exit_time__isnull=False)

        # Search in number plate
        if search_query:
            entries = entries.filter(number_plate__icontains=search_query)

        # Summary-only (exit-based for revenue; exclude zero-amount from paid)
        # Paid metrics: only those with exit today and amount > 0
        paid_qs = entries.filter(
            is_paid=True, total_amount__gt=0, exit_time__isnull=False
        )
        paid_count = paid_qs.count()
        paid_sum = paid_qs.aggregate(total=Sum("total_amount"))["total"] or 0
        # Zero-paid (0 so'm) among paid within window
        unpaid_zero_count = entries.filter(is_paid=True, total_amount=0).count()
        # inside_count should reflect cars that entered within the date range and have no exit
        inside_qs = VehicleEntry.objects.filter(exit_time__isnull=True)
        try:
            if date_from:
                start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                start_datetime = timezone.make_aware(
                    datetime.combine(start_date, datetime.min.time())
                )
                inside_qs = inside_qs.filter(entry_time__gte=start_datetime)
            if date_to:
                end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                end_datetime = timezone.make_aware(
                    datetime.combine(end_date, datetime.max.time())
                )
                inside_qs = inside_qs.filter(entry_time__lte=end_datetime)
        except ValueError:
            pass
        inside_count = inside_qs.count()
        exited_count = entries.filter(exit_time__isnull=False).count()

        filters_snapshot = {
            "date_from": date_from,
            "date_to": date_to,
            "number_plate": number_plate_filter,
            "payment_status": payment_status,
            "entry_status": entry_status,
            "search": search_query,
        }

        # Diagnostics: find resolved printer now (optional)
        resolved_name = None
        default_name = None
        try:
            default_name = win32print.GetDefaultPrinter()
        except Exception:
            default_name = None
        try:
            resolved_name = _resolve_printer_name(printer_name)
        except Exception as _e:
            resolved_name = None

        ok, err = print_stats_receipt(
            printer_name=printer_name,
            date_from=date_from or "",
            date_to=date_to or "",
            filters=filters_snapshot,
            paid_count=paid_count,
            paid_sum=paid_sum,
            unpaid_zero_count=unpaid_zero_count,
            inside_count=inside_count,
            exited_count=exited_count,
        )

        return JsonResponse(
            {
                "status": "ok" if ok else "error",
                "printed": ok,
                "error": err,
                "diagnostics": {
                    "requested_printer": printer_name,
                    "resolved_printer": resolved_name,
                    "default_printer": default_name,
                },
            }
        )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@require_GET
@login_required
def list_printers(request):
    """Return installed printers, default printer, and resolved name (optional)."""
    try:
        preferred = request.GET.get("printer_name")
        installed = [
            p[2]
            for p in win32print.EnumPrinters(
                win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
            )
        ]
        try:
            default_name = win32print.GetDefaultPrinter()
        except Exception:
            default_name = None
        resolved = None
        error = None
        try:
            resolved = _resolve_printer_name(preferred)
        except Exception as ex:
            error = str(ex)
        return JsonResponse(
            {
                "status": "ok",
                "installed": installed,
                "default": default_name,
                "resolved": resolved,
                "error": error,
            }
        )
    except Exception as e:
        return JsonResponse({"status": "error", "error": str(e)}, status=500)


def export_detailed_entries_xls(entries, date_str=None):
    """Export detailed entries to Excel format using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Batafsil yozuvlar"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Calculate daily statistics
        total_entries = len(entries)
        free_exits = sum(
            1
            for entry in entries
            if entry.exit_time
            and (entry.total_amount == 0 or entry.total_amount is None)
        )
        paid_exits = sum(
            1
            for entry in entries
            if entry.exit_time and entry.total_amount and entry.total_amount > 0
        )
        total_revenue = sum(
            entry.total_amount or 0
            for entry in entries
            if entry.exit_time and entry.total_amount
        )
        inside_count = sum(1 for entry in entries if not entry.exit_time)
        exited_count = sum(1 for entry in entries if entry.exit_time)

        # Add title and date
        ws["A1"] = "SMART AUTOPARK - Batafsil yozuvlar"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:H1")

        if date_str:
            ws["A2"] = f"Sana: {date_str}"
            ws["A2"].font = Font(bold=True, size=12)
            ws.merge_cells("A2:H2")

        # Add statistics section
        stats_start_row = 4
        ws[f"A{stats_start_row}"] = "KUNLIK STATISTIKA"
        ws[f"A{stats_start_row}"].font = Font(bold=True, size=14)
        ws[f"A{stats_start_row}"].fill = PatternFill(
            start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
        )
        ws.merge_cells(f"A{stats_start_row}:H{stats_start_row}")

        # Statistics rows
        stats_data = [
            ("Jami kirishlar:", total_entries),
            ("Bepul chiqishlar (0 so'm):", free_exits),
            ("To'lovli chiqishlar:", paid_exits),
            ("Jami tushum:", f"{total_revenue:,} so'm"),
            ("Ichkarida qolganlar:", inside_count),
            ("Chiqib ketganlar:", exited_count),
        ]

        for i, (label, value) in enumerate(stats_data, start=stats_start_row + 1):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"B{i}"].font = Font(bold=True)
            if "tushum" in label.lower():
                ws[f"B{i}"].font = Font(bold=True, color="008000")  # Green for revenue

        # Add empty row before table
        table_start_row = stats_start_row + len(stats_data) + 2

        # Add table headers
        headers = [
            "Avtomobil raqami",
            "Kirish vaqti",
            "Chiqish vaqti",
            "Davomiyligi",
            "To'lov miqdori (so'm)",
            "To'lov holati",
            "Holat",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # Add data rows
        for row_idx, entry in enumerate(entries, start=table_start_row + 1):
            # Calculate duration
            if entry.exit_time:
                duration = entry.exit_time - entry.entry_time
                duration_hours = duration.total_seconds() / 3600
                duration_str = f"{duration_hours:.1f} soat"
            else:
                duration_str = "Ichkarida"

            # Format times
            entry_time = (
                entry.entry_time.strftime("%Y-%m-%d %H:%M") if entry.entry_time else ""
            )
            exit_time = (
                entry.exit_time.strftime("%Y-%m-%d %H:%M") if entry.exit_time else ""
            )

            # Status
            payment_status = "To'langan" if entry.is_paid else "To'lanmagan"
            entry_status = "Ichkarida" if not entry.exit_time else "Chiqib ketgan"

            # Add row data
            row_data = [
                entry.number_plate,
                entry_time,
                exit_time,
                duration_str,
                entry.total_amount or 0,
                payment_status,
                entry_status,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = center_alignment

                # Color coding for payment status
                if col == 6:  # Payment status column
                    if value == "To'langan":
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                    else:
                        cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                        )

                # Color coding for entry status
                if col == 7:  # Entry status column
                    if value == "Ichkarida":
                        cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )
                    else:
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0

            for row in range(1, table_start_row + len(entries) + 1):
                cell_value = ws[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set minimum width and add some padding
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        filename = f"batafsil_yozuvlar_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        if date_str:
            filename = f"batafsil_yozuvlar_{date_str.replace('-', '')}_{timezone.now().strftime('%H%M%S')}.xlsx"

        response["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@require_POST
@login_required
def get_uuid_info(request):
    """UUID bo'yicha ma'lumotlarni olish (to'lov qilmasdan)"""
    try:
        data = json.loads(request.body)
        uuid_code = data.get("uuid")

        if not uuid_code:
            return JsonResponse(
                {"success": False, "error": "UUID kiritish majburiy"}, status=400
            )

        # UUID formatini tekshirish
        uuid_code = uuid_code.strip()  # Bo'shliqlarni olib tashlash

        if len(uuid_code) != 10:
            return JsonResponse(
                {
                    "success": False,
                    "error": f"UUID uzunligi noto'g'ri. Kiritilgan: {len(uuid_code)} ta belgi, kerak: 10 ta belgi",
                },
                status=400,
            )

        if not all(c in "0123456789abcdefABCDEF" for c in uuid_code):
            invalid_chars = [c for c in uuid_code if c not in "0123456789abcdefABCDEF"]
            return JsonResponse(
                {
                    "success": False,
                    "error": f"UUID da noto'g'ri belgilar mavjud: {invalid_chars}",
                },
                status=400,
            )

        # Bazadan UUID bo'yicha yozuvni topish
        try:
            entry = VehicleEntry.objects.get(uuid=uuid_code)
        except VehicleEntry.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": "Bu UUID bilan yozuv topilmadi"}, status=404
            )

        # Agar avtomobil hali chiqib ketmagan bo'lsa, avtomatik chiqish vaqtini belgilash va summani hisoblash
        if not entry.exit_time:
            # Hozirgi vaqtni chiqish vaqti sifatida belgilash
            current_time = timezone.now()
            entry.exit_time = current_time
            # To'lov miqdorini hisoblash va saqlash
            entry.total_amount = entry.calculate_amount()
            entry.save()

        # Ma'lumotlarni qaytarish
        return JsonResponse(
            {
                "success": True,
                "data": {
                    "entry_id": entry.id,
                    "number_plate": entry.number_plate,
                    "entry_time": entry.entry_time.strftime("%Y-%m-%d %H:%M"),
                    "exit_time": entry.exit_time.strftime("%Y-%m-%d %H:%M")
                    if entry.exit_time
                    else None,
                    "total_amount": entry.total_amount or 0,
                    "duration_hours": (
                        entry.exit_time - entry.entry_time
                    ).total_seconds()
                    / 3600
                    if entry.exit_time
                    else 0,
                    "uuid": entry.uuid,
                    "is_paid": entry.is_paid,
                },
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@csrf_exempt
@require_POST
@login_required
def process_uuid_payment(request):
    """UUID orqali to'lov qilish"""
    try:
        data = json.loads(request.body)
        uuid_code = data.get("uuid")

        if not uuid_code:
            return JsonResponse(
                {"success": False, "error": "UUID kiritish majburiy"}, status=400
            )

        # UUID formatini tekshirish
        uuid_code = uuid_code.strip()  # Bo'shliqlarni olib tashlash

        if len(uuid_code) != 10:
            return JsonResponse(
                {
                    "success": False,
                    "error": f"UUID uzunligi noto'g'ri. Kiritilgan: {len(uuid_code)} ta belgi, kerak: 10 ta belgi",
                },
                status=400,
            )

        if not all(c in "0123456789abcdefABCDEF" for c in uuid_code):
            invalid_chars = [c for c in uuid_code if c not in "0123456789abcdefABCDEF"]
            return JsonResponse(
                {
                    "success": False,
                    "error": f"UUID da noto'g'ri belgilar mavjud: {invalid_chars}",
                },
                status=400,
            )

        # Bazadan UUID bo'yicha yozuvni topish
        try:
            entry = VehicleEntry.objects.get(uuid=uuid_code)
        except VehicleEntry.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": "Bu UUID bilan yozuv topilmadi"}, status=404
            )

        # Agar avtomobil allaqachon chiqib ketgan bo'lsa
        if entry.exit_time:
            return JsonResponse(
                {
                    "success": False,
                    "error": f"Avtomobil {entry.number_plate} allaqachon chiqib ketgan",
                },
                status=400,
            )

        # Chiqish vaqtini bugungi sana bilan belgilash
        current_time = timezone.now()
        entry.exit_time = current_time

        # To'lov miqdorini hisoblash
        entry.total_amount = entry.calculate_amount()

        # To'lov qilingan deb belgilash
        entry.is_paid = True

        entry.save()

        # Oddiy chekni chiqarish (consumers.py dagi funksiyadan foydalanish)
        try:
            from .consumers import HomeConsumer

            consumer = HomeConsumer()
            print_result = consumer.print_receipt_to_xprinter_sync(entry.id)
            # Chek chiqarish natijasini log qilish (ixtiyoriy)
            if not print_result.get("success"):
                import logging

                logger = logging.getLogger(__name__)
                logger.warning(f"Chek chiqarishda xatolik: {print_result.get('error')}")
        except Exception as e:
            # Chek chiqarishda xatolik bo'lsa ham to'lov saqlanadi
            import logging

            logger = logging.getLogger(__name__)
            logger.error(f"Chek chiqarishda xatolik: {str(e)}")

        return JsonResponse(
            {
                "success": True,
                "message": "To'lov muvaffaqiyatli amalga oshirildi",
                "data": {
                    "entry_id": entry.id,
                    "number_plate": entry.number_plate,
                    "entry_time": entry.entry_time.strftime("%Y-%m-%d %H:%M"),
                    "exit_time": entry.exit_time.strftime("%Y-%m-%d %H:%M"),
                    "total_amount": entry.total_amount,
                    "duration_hours": (
                        entry.exit_time - entry.entry_time
                    ).total_seconds()
                    / 3600,
                },
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@require_GET
def delete_final_20_days_entries(request):
    """Delete entries that are 20 days old"""
    try:
        with transaction.atomic():
            cutoff_date = timezone.now() - timedelta(days=20)

            VehicleEntry.objects.filter(entry_time__lt=cutoff_date).delete()
            messages.success(request, "20 kunlik yozuvlar o'chirildi")
            return redirect("home")
    except Exception as e:
        messages.error(request, "Xatolik yuz berdi: " + str(e))
        return redirect("home")


class MarkErrorView(View):
    def post(self, request):
        try:
            import json
            data = json.loads(request.body)
            entry_id = data.get('entry_id')
            error_message = data.get('error_message')
            
            if not entry_id or not error_message:
                return JsonResponse({
                    'success': False,
                    'error': 'entry_id va error_message kerak'
                }, status=400)
            
            # Entry ni topish
            try:
                entry = VehicleEntry.objects.get(id=entry_id)
            except VehicleEntry.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Entry topilmadi'
                }, status=404)
            
            # Xatolikni belgilash
            entry.is_error = True
            entry.error_massage = error_message
            entry.save()
            
            
            
            
            return JsonResponse({
                'success': True,
                'message': 'Xatolik muvaffaqiyatli belgilandi',
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'JSON format xatosi'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)